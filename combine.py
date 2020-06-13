import os
import json
from openpyxl import load_workbook
from openpyxl import Workbook
from netflix_movies.uNoGS_movie import uNoGS_movie_method as gs
from netflix_movies.unogsNG_movie import unogsNG_movie as ng


ABSPATH = os.path.abspath(__file__)
# print(ABSPATH)
BASEPATH, current_filename = os.path.split(ABSPATH)
# print(BASEPATH)


gs_final = gs()
print(len(gs_final))
print(gs_final)

ng_final = ng()
print(len(ng_final))
print(ng_final)

raw_rk_final = gs_final + ng_final
print(raw_rk_final)
print(len(raw_rk_final))
rk_final = [list(i) for i in set(map(tuple, raw_rk_final))]
print(rk_final)

wb = Workbook()
excel_path = BASEPATH + "\\"
print(excel_path)
print(os.listdir(excel_path))
if "movies.xlsx" in os.listdir(excel_path):
    wb = load_workbook(excel_path + "movies.xlsx")
    print("excel file already there")
else:
    wb.save(excel_path + "movies.xlsx")
    print("new created and then loaded")
    wb = load_workbook(excel_path + "movies.xlsx")

print(wb.sheetnames)
sheet = wb.active
print(sheet)
for i in range(len(rk_final)):
    print(rk_final[i])
    data = [rk_final[i]]
    for row in data:
        sheet.append(row)
wb.save(excel_path + "movies.xlsx")
